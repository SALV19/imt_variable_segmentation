from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import (
    RichTextProperties,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    CharacterProperties,
    Font as PlotFont,
)
from openpyxl.styles import Font
from openpyxl.chart.title import Title

from utils.get_singular_points import get_singular_points  # pylint: disable=E0401

measure = {
    "IRI": "(m/km)",
    "CF": "",
    "Deflexión máxima": "(mm)",
    "Agriet. fatiga": "(%)",
    "Agriet. long": "(%)",
    "Agriet. transv": "(%)",
    "PR": "(mm)",
    "TDPA": "",
}

static = ["TDPA"]


def generate_sheet(wb: Workbook, title: str, generated_data):
    measurements = generated_data["file_data"]
    slopes = generated_data["slopes"]
    singularities = []
    if "abnormalities" in generated_data:
        singularities = generated_data["abnormalities"]

    slope_values: map = map(
        lambda item: list(map(lambda value: item[value], item)), slopes
    )

    ws = wb.create_sheet(title)

    ws["A1"] = "ID"
    ws["B1"] = measurements["id"]

    ws["A2"] = "Inicio"
    ws["B2"] = "Fin"
    ws["C2"] = title

    ws["E2"] = "Inicio Segmento"
    ws["F2"] = "Fin Segmento"
    ws["G2"] = title
    ws["I2"] = "Puntos Singulares"

    length = len(measurements["measurements"]) + 2

    # Medidas generales
    for idx, measurement in enumerate(measurements["measurements"]):
        row_val = idx + 3
        end_measurement = measurement + measurements["distance"]
        measurement_cell = ws.cell(row=row_val, column=1, value=measurement)
        end_measurement_cell = ws.cell(row=row_val, column=2, value=end_measurement)
        measurement_cell.number_format = '0"+"000'
        end_measurement_cell.number_format = '0"+"000'
        value_cell = ws.cell(row=row_val, column=3, value=measurements["values"][idx])
        value_cell.number_format = "#,##0.00"

    slope_values = list(slope_values)

    # Gráfica 2: pendiente
    for idx, value in enumerate(slope_values):
        row_val = idx + 3
        ws.cell(row=row_val, column=5, value=value[0])
        ws.cell(row=row_val, column=6, value=value[1])
        ws.cell(row=row_val, column=7, value=value[2])
        cell = ws.cell(row=row_val, column=8, value=value[2])

        cell.font = Font(color="FFFFFF")

    # Puntos singulares
    idx = 0
    singularities_length = len(singularities)
    for i in range(len(measurements["measurements"])):
        if idx >= singularities_length:
            break

        if measurements["measurements"][i] == singularities[idx]["x"]:
            ws.cell(row=idx + 3, column=9, value=singularities[idx]["x"])
            ws.cell(row=idx + 3, column=10, value=singularities[idx]["y"])
            idx += 1

    c1 = ScatterChart()

    text = Text()
    char_props = CharacterProperties(
        sz=1400,
        solidFill="999999",
        latin=PlotFont(typeface="Roboto"),
    )

    if title in static:
        paragraph = Paragraph(
            pPr=ParagraphProperties(defRPr=char_props),
            r=[RegularTextRun(t=f"Segmentación para {title}")],
        )

        rich_text = RichText(p=[paragraph])
        text = Text(rich=rich_text)

    else:
        measurement_values = Reference(ws, min_col=1, min_row=3, max_row=length)

        iri = Reference(ws, min_col=3, min_row=3, max_row=length)

        iri_series = Series(values=iri, xvalues=measurement_values, title=title)
        iri_series.graphicalProperties.line.solidFill = "0F9ED5 "
        iri_series.smooth = False

        c1.series.append(iri_series)

        paragraph = Paragraph(
            pPr=ParagraphProperties(defRPr=char_props),
            r=[RegularTextRun(t=f"Segmentación dinámica para {title}")],
        )

        rich_text = RichText(p=[paragraph])

        text = Text(rich=rich_text)

    c1.title = Title(text)
    c1.style = 2
    c1.y_axis.title = title + " " + measure[title]
    c1.x_axis.tickLblPos = "nextTo"
    c1.x_axis.scaling.min = measurements["measurements"][0]
    c1.x_axis.scaling.max = measurements["measurements"][-1]
    c1.x_axis.scaling.orientation = "minMax"
    c1.x_axis.number_format = '0"+"000'
    c1.roundedCorners = False

    c1.x_axis.majorUnit = round(
        (measurements["measurements"][-1] - measurements["measurements"][0]) / 10
    )

    if singularities_length:
        singular_points = get_singular_points(ws, singularities_length)

        c1.series.append(singular_points)

    for i in range(len(slope_values)):
        x_segmentation_series = Reference(ws, min_col=5, max_col=6, min_row=i + 3)
        y_segmentation_series = Reference(ws, min_col=7, max_col=8, min_row=i + 3)
        segmentation_series: Series = Series(
            values=y_segmentation_series,
            xvalues=x_segmentation_series,
            title=(lambda i: "Segmentos" if i == 0 else "Other")(i),
        )
        segmentation_series.graphicalProperties.line.solidFill = "000000"

        c1.series.append(segmentation_series)

    c1.width = 30
    c1.height = 15

    c1.x_axis.delete = False

    c1.layout = Layout(
        manualLayout=ManualLayout(
            x=-0.02,
            y=0.01,
            h=0.9,
            w=0.78,
        )
    )

    c1.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(
            anchor="ctr",
            anchorCtr="1",
            rot="-2700000",
            spcFirstLastPara="1",
            vertOverflow="ellipsis",
            wrap="square",
        ),
        p=[
            Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties()),
                endParaRPr=CharacterProperties(),
            )
        ],
    )

    # Chart Styling

    # ---- Adjust line thickness and color ----
    if not title in static:
        c1.series[0].graphicalProperties.line.width = 15000

    c1.legend.position = "b"
    c1.legend.layout = Layout(
        manualLayout=ManualLayout(
            x=0.4,
            y=0.95,
        )
    )

    ws.add_chart(c1, "L2")


# Excel Macro to delete all extra labels

# Sub QuitarLabels()
# '
# ' QuitarLabels Macro
# '

# '
# Dim wb As Workbook
# Dim ws As Worksheet
# Dim chart As ChartObject
# Dim legend As LegendEntry

# Set wb = ActiveWorkbook

# For Each ws In wb.Sheets
#     If ws.ChartObjects.Count > 0 Then
#         Set chartObj = ws.ChartObjects(1)
#         If chartObj.chart.HasLegend Then
#             legendCount = chartObj.chart.legend.LegendEntries.Count
#             While legendCount >= 4
#                 chartObj.chart.legend.LegendEntries(4).Delete
#             Wend
#         End If
#     End If
# Next ws

# End Sub
