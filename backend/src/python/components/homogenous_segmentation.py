from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import (
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    CharacterProperties,
    Font as PlotFont,
)
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker

titles = {
    "iri": "IRI",  # (m/km)
    "friccion": "CF",
    "deflexiones": "Deflexión máxima",  # (mm)
    "agrfatiga": "Agriet. fatiga",  # (%)
    "grlong": "Agriet. long",  # (%)
    "grtrans": "Agriet. transv",  # (%)
    "pr": "PR",  # (mm)
    "tdpa": "TDPA",
    "baches": "Baches",
    "desprendimientos": "Desprendimientos",  # (%)
    "macrotextura": "Macrotextura",  # (mm)
    "tipo de pavimento": "Tipo de pavimento",
}


def homogenous_segmentation(wb: Workbook, h_segmentation):
    ws = wb.create_sheet("Segmentos Homogeneos")

    ws.cell(row=2, column=2, value="Inicio")
    ws.cell(row=2, column=3, value="Fin")

    base_col = 4

    for idx, value in enumerate(h_segmentation):
        row_val = idx + 3
        ws.cell(row=row_val, column=2, value=value["start"])
        ws.cell(row=row_val, column=3, value=value["end"])

        counter = base_col
        for key, val in value["parameters"].items():
            # write header once (when idx == 0)
            if idx == 0:
                ws.cell(row=2, column=counter, value=key.capitalize())
            ws.cell(row=row_val, column=counter, value=val)

            counter += 1

    # Create matrix of segments
    parameters = {}

    # inicialización
    for param in h_segmentation[0]["parameters"]:
        parameters[param] = [h_segmentation[0]["start"]]

    # detección de cambios
    for i in range(1, len(h_segmentation)):
        prev = h_segmentation[i - 1]
        curr = h_segmentation[i]

        for param, curr_val in curr["parameters"].items():
            prev_val = prev["parameters"][param]

            if curr_val != prev_val:
                parameters[param].append(curr["start"])

    # cierre al final
    end_global = h_segmentation[-1]["end"]
    for param in parameters:
        parameters[param].append(end_global)

    chart = ScatterChart()

    for idx, (parameter, positions) in enumerate(parameters.items()):
        column = 18 + idx * 2

        for i, position in enumerate(positions):
            position_cell = ws.cell(row=i + 3, column=column, value=position)
            position_cell.font = Font(color="FFFFFF")

            idx_cell = ws.cell(row=i + 3, column=column + 1, value=idx + 1)
            idx_cell.font = Font(color="FFFFFF")

        length = len(parameters[parameter]) + 3

        x_series_references = Reference(ws, min_col=column, min_row=3, max_row=length)
        idx_series_reference = Reference(
            ws, min_col=column + 1, min_row=3, max_row=length
        )

        serie = Series(
            values=idx_series_reference,
            xvalues=x_series_references,
            title=titles[parameter],
        )
        serie.marker = Marker("circle")
        serie.marker.size = 8

        chart.series.append(serie)

    # Chart customization
    chart.style = 2
    chart.y_axis.delete = True
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "nextTo"
    chart.x_axis.scaling.orientation = "minMax"
    chart.x_axis.number_format = '0"+"000'
    chart.roundedCorners = False

    chart.width = 30
    chart.height = 15

    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=-0.02,
            y=0.01,
            h=0.9,
            w=0.78,
        )
    )

    char_props = CharacterProperties(
        sz=1400,
        solidFill="999999",
        latin=PlotFont(typeface="Roboto"),
    )

    paragraph = Paragraph(
        pPr=ParagraphProperties(defRPr=char_props),
        r=[RegularTextRun(t="Segmentación Homogenea")],
    )

    rich_text = RichText(p=[paragraph])
    text = Text(rich=rich_text)
    chart.title = Title(text)

    chart.legend.position = "b"
    chart.legend.layout = Layout(
        manualLayout=ManualLayout(
            x=0.4,
            y=0.95,
        )
    )

    ws.add_chart(chart, "R2")
