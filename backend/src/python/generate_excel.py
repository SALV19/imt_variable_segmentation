from io import BytesIO
from openpyxl import load_workbook
import json
import sys

from components.worksheet_generator import generate_sheet
from components.homogenous_segmentation import (
    homogenous_segmentation,
)  # pylint: disable=E0611

from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

dynamic_segmentation = input()
generated_data = json.loads(dynamic_segmentation)

satic_segmentation = input()
static_data = json.loads(satic_segmentation)

hsegment = input()
h_segmentation = json.loads(hsegment)

wb = load_workbook("./src/python/template.xlsm", keep_vba=True)
virtual_workbook = BytesIO()

titles = {
    "iri": "IRI",  # (m/km)
    "friccion": "CF",
    "deflexiones": "Deflexión máxima",  # (mm)
    "agrfatiga": "Agriet. fatiga",  # (%)
    "grlong": "Agriet. long",  # (%)
    "grtrans": "Agriet. transv",  # (%)
    "pr": "PR",  # (mm)
    "tdpa": "TDPA",
    "baches": "Baches",
    "desprendimientos": "Desprendimientos",  # (%)
    "macrotextura": "Macrotextura",  # (mm)
    "tipo de pavimento": "Tipo de pavimento",
}

for gd in generated_data:
    key = list(gd.keys())[0]
    values = list(gd.values())[0]["generated_data"]
    generate_sheet(wb, titles[key], values)

for sd in static_data:
    key = list(sd.keys())[0]
    values = list(sd.values())[0]
    generate_sheet(wb, titles[key], values)

homogenous_segmentation(wb, h_segmentation)

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

if "Hoja1" in wb.sheetnames:
    del wb["Hoja1"]

wb.save(virtual_workbook)

sys.stdout.buffer.write(virtual_workbook.getvalue())
